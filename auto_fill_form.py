from openpyxl import load_workbook
from openpyxl import utils as xl_u
import time
import re
import os
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as expected_conditions
from selenium.webdriver.common.by import By 
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options

print('import successful')

# You can customize these variables 
excel_filename = 'input.xlsx'
excel_input_sheet = 'Hoja1'
excel_input_path = os.getcwd() + '\\' + excel_filename  
print('Excel path -> ' + excel_input_path)                       

##################################################################

# Define variables
name = []
surname = []
email = []
url = 'http://latinresearch.org/65-2/'

# Setup selenium
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('disable_infobars')
driver = webdriver.Chrome(options=chrome_options)

# Connect to Excel
wb = load_workbook(excel_input_path)
sheet = wb[excel_input_sheet]
num_row = sheet.max_row
print('\n---------------------')
print('Number of row/input: ' + str(num_row))
print('---------------------')

def auto_fill():
    # Loop through each row
    for x in range(num_row):
        try:
            name_value = sheet['A' + str(x+1)].value
            surname_value = sheet['B' + str(x+1)].value
            email_value = sheet['C' + str(x+1)].value
        except Exception as err:
            print('Error getting value -> ' + str(err))

        driver.get(url)
        time.sleep(2)
        input_name = driver.find_element_by_xpath("//input[@name='dk-speakout-first-name']")
        input_surname = driver.find_element_by_xpath("//input[@name='dk-speakout-last-name']")
        input_email = driver.find_element_by_xpath("//input[@name='dk-speakout-email']")

        # Send input value
        try:
            print('Input ' + name_value + ', ' + surname_value + ', ' + email_value)
            input_name.click()
            input_name.send_keys(name_value)
            input_surname.click()
            input_surname.send_keys(surname_value)
            input_email.click()
            input_email.send_keys(email_value)
        except TypeError:
            pass
        except Exception as err:
            print('Error input value -> ' + str(err))

        # Interval
        # driver.find_element_by_xpath("//button[@class='dk-speakout-submit']").click() # CLICK THE BUTTON
        time.sleep(3)
        
        name.append(name_value) 
        surname.append(surname_value)
        email.append(email_value)

auto_fill()
# In case you wanna check
# print('======================')
# print(name)
# print(surname)
# print(email)
print('---------------------')
print('Fill form successful')
# time.sleep(10)
# driver.quit()


